from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import InventoryItem, StockMovement, SensorBuild, SensorComponent, DeployedSensor
from .forms import InventoryItemForm
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Spacer, Image, Table, TableStyle, HRFlowable
from reportlab.lib.styles import ParagraphStyle
from django.conf import settings
import os
from datetime import datetime
from django.utils import timezone
from django.db import transaction
from django.core.paginator import Paginator

from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
import openpyxl

def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_admin)
def add_inventory_item(request):
    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            part_number = form.cleaned_data['part_number']
            total_stock = form.cleaned_data['total_stock']
            
            # Check for existing item with same part number
            existing_item = InventoryItem.objects.filter(part_number=part_number).first()
            
            if existing_item:
                # Update existing item's stock
                existing_item.total_stock += total_stock
                existing_item.save()
                
                # Create stock movement for additional stock
                StockMovement.objects.create(
                    item=existing_item,
                    movement_type='INWARD',
                    quantity=total_stock,
                    purpose='Procurement',
                    remarks='Additional stock added'
                )
            else:
                # Create new item
                item = form.save()
                StockMovement.objects.create(
                    item=item,
                    movement_type='INWARD', 
                    quantity=item.total_stock,
                    purpose='Procurement',
                    remarks='Initial stock added'
                )
            
            return redirect('inventory_list')
        else:
            print("Form errors:", form.errors)
    else:
        form = InventoryItemForm()
    return render(request, 'inventory_item_form.html', {'form': form})

@login_required
def inventory_list(request):
    items = InventoryItem.objects.all()
    raw_materials_count = InventoryItem.objects.filter(category='components').count()
    sensors_count = InventoryItem.objects.filter(category='built_equipment').count()
    return render(request, 'inventory_list.html', {
        'items': items,
        'raw_materials_count': raw_materials_count,
        'sensors_count': sensors_count
    })

@login_required
# @user_passes_test(is_admin)
def stock_movement(request):
    items = InventoryItem.objects.all()
    if request.method == 'POST':
        item_id = request.POST['item']
        movement_type = request.POST['movement_type']
        quantity = int(request.POST['quantity'])
        vendor = request.POST.get('vendor', '')
        purpose = request.POST.get('purpose', '')

        item = InventoryItem.objects.get(id=item_id)
        if movement_type == 'OUTWARD':
            if item.total_stock < quantity:
                return render(request, 'stock_movement_form.html', {'error': 'Insufficient stock!', 'items': items})
            item.total_stock -= quantity
        elif movement_type == 'INWARD':
            item.total_stock += quantity

        item.save()
        StockMovement.objects.create(item=item, movement_type=movement_type, quantity=quantity, vendor=vendor, purpose=purpose)
        return redirect('stock_list')

    return render(request, 'stock_movement_form.html', {'items': items})

@login_required
def stock_list(request):
    # Fetch all stock movements ordered by date
    stock_movements = StockMovement.objects.all().order_by('-date')
    
    # Paginate the stock movements (8 per page)
    paginator = Paginator(stock_movements, 8)
    page_number = request.GET.get('page')  # Get the current page number from the request
    page_obj = paginator.get_page(page_number)  # Get the page object
    
    return render(request, 'stock_list.html', {'page_obj': page_obj})



@login_required
# @user_passes_test(is_admin)
def sensor_build(request):
    raw_materials = InventoryItem.objects.filter(category='components')
    if request.method == 'POST':
        sensor_name = request.POST['sensor_name']
        part_number = request.POST['part_number']
        components = request.POST.getlist('components')
        quantities = request.POST.getlist('quantities')
        number_of_sensors = int(request.POST['number_of_sensors'])

        # Validate stock availability before proceeding
        for component_id, quantity in zip(components, quantities):
            component = InventoryItem.objects.get(id=component_id)
            quantity_used = int(quantity) 
            
            if component.total_stock < quantity_used:
                return render(request, 'sensor_build_form.html', 
                             {'error': f'Insufficient stock for {component.name}! Need {quantity_used}, have {component.total_stock}', 
                              'raw_materials': raw_materials})

        with transaction.atomic():
            # Create the new sensor with the correct stock amount from the beginning
            new_sensor = InventoryItem.objects.create(
                part_number=part_number,
                name=sensor_name,
                category='built_equipment',
                description='',
                total_stock=number_of_sensors  # Set the correct stock immediately
            )
            
            # Force refresh from database to ensure we have the latest data
            new_sensor.refresh_from_db()

            # Create the SensorBuild record
            sensor_build = SensorBuild.objects.create(sensor=new_sensor)

            # Create SensorComponent records and update stock movements
            for component_id, quantity in zip(components, quantities):
                component = InventoryItem.objects.get(id=component_id)
                quantity_used = int(quantity) 

                component.total_stock -= quantity_used
                component.save()
                component.refresh_from_db()

                SensorComponent.objects.create(
                    sensor_build=sensor_build,
                    component=component,
                    quantity_used=int(quantity)
                )

                # Create a StockMovement record for outward movement of raw materials
                StockMovement.objects.create(
                    item=component,
                    movement_type='OUTWARD',
                    quantity=quantity_used,
                    purpose='Production',
                    remarks=f'Used in building {number_of_sensors} {new_sensor.name}'
                )

            # Create a StockMovement record for inward movement of the new sensor
            StockMovement.objects.create(
                item=new_sensor,
                movement_type='INWARD',
                quantity=number_of_sensors,
                purpose='Production',
                remarks=f'{number_of_sensors} sensors built'
            )

        return redirect('inventory_list')

    return render(request, 'sensor_build_form.html', {'raw_materials': raw_materials})

@login_required
def sensor_detail(request, sensor_id):
    sensor = get_object_or_404(InventoryItem, id=sensor_id, category='built_equipment')
    sensor_builds = SensorBuild.objects.filter(sensor=sensor).prefetch_related('components__component')
    return render(request, 'sensor_details.html', {'sensor': sensor, 'sensor_builds': sensor_builds})

@login_required
def generate_pdf(request):
    items = InventoryItem.objects.all()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.pdf"'
    
    # Set up the document with better margins
    doc = SimpleDocTemplate(
        response, 
        pagesize=letter,
        rightMargin=50, 
        leftMargin=50,
        topMargin=50, 
        bottomMargin=50
    )
    
    elements = []
    
    # Define colors for a professional look
    brand_color = colors.HexColor('#7c3aed')  # Purple from your secondary color
    accent_color = colors.HexColor('#3b82f6')  # Blue from your primary color
    text_color = colors.HexColor('#1e293b')   # Dark slate
    muted_color = colors.HexColor('#64748b')  # Gray
    light_color = colors.HexColor('#f8fafc')  # Very light gray
    border_color = colors.HexColor('#e2e8f0') # Light gray for borders
    
    # Create a title with company branding
    styles = getSampleStyleSheet()
    
    # Company name and logo section
    company_style = ParagraphStyle(
        'Company',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=brand_color,
        leading=26,
        spaceAfter=0,
    )
    
    # Add a table for the header (logo + title)
    header_data = [[
        Paragraph('<b>StockForge</b>', company_style),
        Paragraph(f'<font color="#64748b" size="9">Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}</font>', 
                 styles['Normal'])
    ]]
    
    header_table = Table(header_data, colWidths=[doc.width/2.0]*2)
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header_table)
    
    # Add an optional logo here
    # logo_path = os.path.join(settings.STATIC_ROOT, 'img/logo.png')
    # if os.path.exists(logo_path):
    #     logo = Image(logo_path, width=1.5*inch, height=0.5*inch)
    #     elements.append(logo)
    
    # Add horizontal line after header
    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(
        width="100%",
        thickness=1,
        color=border_color,
        spaceAfter=0,
        spaceBefore=0,
        hAlign='CENTER',
    ))
    elements.append(Spacer(1, 20))
    
    # Report Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=text_color,
        alignment=1,  # Center alignment
        spaceAfter=10,
    )
    elements.append(Paragraph("Inventory Report", title_style))
    
    # Add report summary section
    components_count = InventoryItem.objects.filter(category='components').count()
    equipment_count = InventoryItem.objects.filter(category='built_equipment').count()
    
    summary_data = [
        ['Total Inventory Items:', f"{items.count()} items"],
        ['Components:', f"{components_count} items"],
        ['Built Equipment:', f"{equipment_count} items"],
        ['Generated On:', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['Generated By:', request.user.get_full_name() or request.user.username],
    ]
    
    summary_table = Table(summary_data, colWidths=[150, 200])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), brand_color),
        ('TEXTCOLOR', (1, 0), (1, -1), text_color),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    # Wrap the summary table in a container with background
    summary_background = Table([[summary_table]], colWidths=[doc.width])
    summary_background.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), light_color),
        ('ROUNDEDCORNERS', [6, 6, 6, 6]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    elements.append(Spacer(1, 15))
    elements.append(summary_background)
    elements.append(Spacer(1, 20))
    
    # Table header
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=text_color,
        spaceAfter=8,
        spaceBefore=15,
    )
    elements.append(Paragraph("Inventory Items", section_style))
    elements.append(Spacer(1, 5))
    
    # Create table data with more detailed information
    data = [
        ['Part Number', 'Item Name', 'Category', 'Stock Level', 'Status']
    ]
    
    for item in items:
        # Determine stock status
        if item.total_stock <= 0:
            stock_status = "Out of Stock"
            status_color = colors.HexColor('#ef4444')  # Red
        elif item.total_stock < 5:
            stock_status = "Low Stock"
            status_color = colors.HexColor('#f59e0b')  # Orange
        elif item.total_stock < 20:
            stock_status = "Medium Stock"
            status_color = colors.HexColor('#10b981')  # Green
        else:
            stock_status = "Well Stocked"
            status_color = colors.HexColor('#10b981')  # Green
        
        data.append([
            item.part_number,
            item.name,
            item.get_category_display(),
            str(item.total_stock),
            stock_status
        ])
    
    # Calculate column widths - allocate proportionally
    col_widths = [80, 150, 100, 70, 80]
    
    # Create table with professional styling
    table = Table(data, repeatRows=1, colWidths=col_widths)
    
    # Complex table styling for a professional look
    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0,0), (-1,0), brand_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        
        # Body
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('TEXTCOLOR', (0,1), (-1,-1), text_color),
        
        # Alignment
        ('ALIGN', (0,0), (0,-1), 'LEFT'),  # Part number
        ('ALIGN', (1,0), (1,-1), 'LEFT'),  # Name
        ('ALIGN', (2,0), (2,-1), 'LEFT'),  # Category
        ('ALIGN', (3,0), (3,-1), 'CENTER'),  # Stock level
        ('ALIGN', (4,0), (4,-1), 'CENTER'),  # Status
        
        # Grid lines
        ('LINEBELOW', (0,0), (-1,0), 1, colors.white),
        ('LINEBELOW', (0,1), (-1,-1), 0.5, border_color),
        ('LINEAFTER', (0,0), (-2,-1), 0.5, border_color),
        
        # Padding
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        
        # Alternate row background
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_color]),
    ])
    
    # Apply conditional formatting for stock status
    for i in range(1, len(data)):
        status = data[i][4]
        if status == "Out of Stock":
            table_style.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#ef4444'))
            table_style.add('BACKGROUND', (4, i), (4, i), colors.HexColor('#fee2e2'))
        elif status == "Low Stock":
            table_style.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#f59e0b'))
            table_style.add('BACKGROUND', (4, i), (4, i), colors.HexColor('#fef3c7'))
        else:
            table_style.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#10b981'))
            table_style.add('BACKGROUND', (4, i), (4, i), colors.HexColor('#d1fae5'))
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Add note section
    elements.append(Spacer(1, 30))
    note_style = ParagraphStyle(
        'Note',
        parent=styles['Normal'],
        fontSize=9,
        textColor=muted_color,
        alignment=0,
        leftIndent=0,
    )
    elements.append(Paragraph("<b>Notes:</b> This is an automatically generated inventory report. Low stock items may require restocking.", note_style))
    
    # Add footer with page numbers
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(muted_color)
        
        # Add footer line
        canvas.line(50, 40, letter[0] - 50, 40)
        canvas.setLineWidth(0.5)
        canvas.setStrokeColor(border_color)
        
        # Add company info and page numbers
        footer_text = "StockForge Inventory Management System"
        page_text = f"Page {canvas.getPageNumber()}"
        
        canvas.drawString(50, 25, footer_text)
        canvas.drawRightString(letter[0] - 50, 25, page_text)
        canvas.restoreState()
    
    # Build the document with page numbers
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    return response

@login_required
def deployed_sensors(request):
    if request.method == 'POST':
        sensor_id = request.POST['sensor']
        company = request.POST['company']
        deployment_date = request.POST['deployment_date']
        expected_return_date = request.POST.get('expected_return_date')
        quantity = int(request.POST['quantity'])
        
        sensor = get_object_or_404(InventoryItem, id=sensor_id, category='built_equipment')
        
        if sensor.total_stock < quantity:
            # Get the appropriate deployments based on user role
            if request.user.is_superuser:
                deployments = DeployedSensor.objects.all()
            else:
                deployments = DeployedSensor.objects.filter(deployed_by=request.user)
                
            return render(request, 'deployed_sensors.html', {
                'error': f'Insufficient stock for {sensor.name}! Need {quantity}, have {sensor.total_stock}',
                'sensors': InventoryItem.objects.filter(category='built_equipment'),
                'deployments': deployments
            })
        
        # Create single deployment record with quantity and associate with current user
        deployment = DeployedSensor.objects.create(
            sensor=sensor,
            company=company,
            deployment_date=deployment_date,
            expected_return_date=expected_return_date,
            status='DEPLOYED',
            quantity=quantity,
            deployed_by=request.user  # Associate with current user
        )
        
        # Update stock and create movement
        sensor.total_stock -= quantity
        sensor.save()
        
        StockMovement.objects.create(
            item=sensor,
            movement_type='OUTWARD',
            quantity=quantity,
            purpose='Deployment',
            remarks=f'{quantity} sensors deployed to {company}'
        )
        
        return redirect('deployed_sensors')
    
    # Filter deployments based on user role
    if request.user.is_superuser:
        deployments = DeployedSensor.objects.all()
    else:
        deployments = DeployedSensor.objects.filter(deployed_by=request.user)
    
    return render(request, 'deployed_sensors.html', {
        'sensors': InventoryItem.objects.filter(category='built_equipment'),
        'deployments': deployments
    })

@login_required
@user_passes_test(is_admin)
def update_deployment_status(request, pk):
    deployment = get_object_or_404(DeployedSensor, pk=pk)
    previous_status = deployment.status
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        
        if new_status != previous_status:
            sensor = deployment.sensor
            
            if new_status == 'RETURNED':
                if previous_status in ['DEPLOYED', 'SOLD']:
                    sensor.total_stock += deployment.quantity
                    sensor.save()
                    deployment.actual_return_date = timezone.now()
                    
                    StockMovement.objects.create(
                        item=sensor,
                        movement_type='INWARD',
                        quantity=deployment.quantity,
                        purpose='Return',
                        remarks=f'Returned {deployment.quantity} from {deployment.company}'
                    )

                
            elif new_status in ['DEPLOYED', 'SOLD']:
                if previous_status == 'RETURNED':
                    if sensor.total_stock >= deployment.quantity:
                        sensor.total_stock -= deployment.quantity
                        sensor.save()
                        
                        movement_type = 'Deployment' if new_status == 'DEPLOYED' else 'Sale'
                        StockMovement.objects.create(
                            item=sensor,
                            movement_type='OUTWARD',
                            quantity=deployment.quantity,
                            purpose=movement_type,
                            remarks=f'{movement_type} of {deployment.quantity} to {deployment.company}'
                        )
                    else:
                        return redirect('deployed_sensors')
                        
            deployment.status = new_status
            deployment.save()
    
    return redirect('deployed_sensors')



# excel report generation method

@login_required
def export_to_excel(request):
    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventory Report"

    # Add header row
    headers = ["Part Number", "Item Name", "Category", "Stock Level", "Status"]
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Fetch inventory data
    items = InventoryItem.objects.all()

    # Add data rows
    for row_num, item in enumerate(items, 2):
        # Determine stock status
        if item.total_stock <= 0:
            stock_status = "Out of Stock"
        elif item.total_stock < 5:
            stock_status = "Low Stock"
        elif item.total_stock < 20:
            stock_status = "Medium Stock"
        else:
            stock_status = "Well Stocked"

        # Add data to the row
        sheet.cell(row=row_num, column=1).value = item.part_number
        sheet.cell(row=row_num, column=2).value = item.name
        sheet.cell(row=row_num, column=3).value = item.get_category_display()
        sheet.cell(row=row_num, column=4).value = item.total_stock
        sheet.cell(row=row_num, column=5).value = stock_status

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="inventory_report.xlsx"'
    workbook.save(response)
    return response